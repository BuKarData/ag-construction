import csv
import hashlib
import io
import json
from datetime import date, datetime

from django.conf import settings
from django.http import HttpResponse

from .models import Inwestycja, Oferta


# ── helpers ──────────────────────────────────────────────────────────────────

def _md5(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


def _open_data_headers(response, etag: str) -> HttpResponse:
    response["Access-Control-Allow-Origin"] = "*"
    response["Cache-Control"] = "public, max-age=3600"
    response["ETag"] = f'"{etag}"'
    response["Last-Modified"] = datetime.utcnow().strftime("%a, %d %b %Y %H:%M:%S GMT")
    response["Link"] = '<https://creativecommons.org/publicdomain/zero/1.0/>; rel="license"'
    return response


def _get_oferty():
    return (
        Oferta.objects
        .select_related("inwestycja", "rodzaj_lokalu")
        .prefetch_related("ceny", "pomieszczenia", "rabaty", "swiadczenia")
        .filter(status="dostepne")
        .order_by("inwestycja__id", "numer_lokalu")
    )


def _aktualna_cena(oferta):
    ceny = list(oferta.ceny.all())
    return ceny[-1].kwota if ceny else None


def _cena_m2(oferta):
    cena = _aktualna_cena(oferta)
    if cena and oferta.metraz:
        return round(float(cena) / float(oferta.metraz), 2)
    return None


def _cena_pomieszczen(oferta):
    total = sum(float(p.cena) for p in oferta.pomieszczenia.all() if p.cena)
    return round(total, 2) if total else None


def _cena_swiadczen(oferta):
    total = sum(float(s.kwota) for s in oferta.swiadczenia.all() if s.kwota)
    return round(total, 2) if total else None


def _base_row(oferta) -> dict:
    nip = getattr(settings, "COMPANY_NIP", "")
    regon = getattr(settings, "COMPANY_REGON", "")
    nazwa = getattr(settings, "COMPANY_NAME", "AG Construction")
    adres = getattr(settings, "COMPANY_ADDRESS", "")
    telefon = getattr(settings, "COMPANY_PHONE", "")
    email = getattr(settings, "COMPANY_EMAIL", "")

    inv = oferta.inwestycja
    cena = _aktualna_cena(oferta)

    return {
        "nip": nip,
        "regon": regon,
        "nazwa_firmy": nazwa,
        "adres_firmy": adres,
        "telefon": telefon,
        "email": email,
        "id_przedsiewziecia": inv.unikalny_identyfikator_przedsiewziecia or f"inv-{inv.pk}",
        "nazwa_przedsiewziecia": inv.nazwa,
        "adres_przedsiewziecia": inv.adres,
        "numer_lokalu": oferta.numer_lokalu or "",
        "numer_oferty": oferta.numer_oferty or "",
        "rodzaj_lokalu": str(oferta.rodzaj_lokalu) if oferta.rodzaj_lokalu else "",
        "powierzchnia_uzytkowa_m2": str(oferta.metraz) if oferta.metraz else "",
        "liczba_pokoi": oferta.pokoje,
        "cena_lokalu_brutto_pln": str(cena) if cena else "",
        "cena_za_m2_brutto_pln": str(_cena_m2(oferta)) if _cena_m2(oferta) else "",
        "cena_pomieszczen_przynaleznych_pln": str(_cena_pomieszczen(oferta)) or "",
        "inne_swiadczenia_pieniezne_pln": str(_cena_swiadczen(oferta)) or "",
        "data_aktualizacji": date.today().isoformat(),
    }


def _flat_rows():
    """Flatten all offers + dynamic pomieszczenia/rabaty/swiadczenia columns."""
    oferty = _get_oferty()
    rows = []
    max_pom = max_rab = max_sw = 0

    base_rows = []
    for o in oferty:
        row = _base_row(o)
        pom = list(o.pomieszczenia.all())
        rab = list(o.rabaty.all())
        sw = list(o.swiadczenia.all())
        for i, p in enumerate(pom, 1):
            row[f"pomieszczenie_{i}_nazwa"] = p.nazwa
            row[f"pomieszczenie_{i}_powierzchnia_m2"] = str(p.powierzchnia) if p.powierzchnia else ""
            row[f"pomieszczenie_{i}_cena_pln"] = str(p.cena) if p.cena else ""
        for i, r in enumerate(rab, 1):
            row[f"rabat_{i}_nazwa"] = r.nazwa
            row[f"rabat_{i}_wartosc"] = str(r.wartosc)
            row[f"rabat_{i}_typ"] = r.typ
        for i, s in enumerate(sw, 1):
            row[f"swiadczenie_{i}_nazwa"] = s.nazwa
            row[f"swiadczenie_{i}_kwota_pln"] = str(s.kwota)
        max_pom = max(max_pom, len(pom))
        max_rab = max(max_rab, len(rab))
        max_sw = max(max_sw, len(sw))
        base_rows.append(row)

    # Build consistent fieldnames
    fixed = list(_base_row.__code__.co_consts)  # unused — build manually
    fixed_fields = [
        "nip", "regon", "nazwa_firmy", "adres_firmy", "telefon", "email",
        "id_przedsiewziecia", "nazwa_przedsiewziecia", "adres_przedsiewziecia",
        "numer_lokalu", "numer_oferty", "rodzaj_lokalu",
        "powierzchnia_uzytkowa_m2", "liczba_pokoi",
        "cena_lokalu_brutto_pln", "cena_za_m2_brutto_pln",
        "cena_pomieszczen_przynaleznych_pln", "inne_swiadczenia_pieniezne_pln",
        "data_aktualizacji",
    ]
    for i in range(1, max_pom + 1):
        fixed_fields += [f"pomieszczenie_{i}_nazwa", f"pomieszczenie_{i}_powierzchnia_m2", f"pomieszczenie_{i}_cena_pln"]
    for i in range(1, max_rab + 1):
        fixed_fields += [f"rabat_{i}_nazwa", f"rabat_{i}_wartosc", f"rabat_{i}_typ"]
    for i in range(1, max_sw + 1):
        fixed_fields += [f"swiadczenie_{i}_nazwa", f"swiadczenie_{i}_kwota_pln"]

    # Fill missing keys with ""
    for row in base_rows:
        for f in fixed_fields:
            row.setdefault(f, "")
        rows.append({f: row[f] for f in fixed_fields})

    return fixed_fields, rows


# ── CSV ───────────────────────────────────────────────────────────────────────

def generate_csv() -> bytes:
    fields, rows = _flat_rows()
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=fields, delimiter=";", lineterminator="\r\n")
    writer.writeheader()
    writer.writerows(rows)
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def view_csv(request):
    data = generate_csv()
    etag = _md5(data)
    today = date.today().strftime("%Y%m%d")
    resp = HttpResponse(data, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="ag_construction_oferty_{today}.csv"'
    return _open_data_headers(resp, etag)


# ── XLSX ──────────────────────────────────────────────────────────────────────

def generate_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    fields, rows = _flat_rows()
    wb = Workbook()
    ws = wb.active
    ws.title = "Raport ofert"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D1B2A")

    for col_idx, field in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(field) + 2, 14)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[field])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def view_xlsx(request):
    data = generate_xlsx()
    etag = _md5(data)
    today = date.today().strftime("%Y%m%d")
    resp = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="ag_construction_oferty_{today}.xlsx"'
    return _open_data_headers(resp, etag)


# ── JSON-LD ───────────────────────────────────────────────────────────────────

def generate_jsonld() -> dict:
    nip = getattr(settings, "COMPANY_NIP", "")
    regon = getattr(settings, "COMPANY_REGON", "")
    nazwa = getattr(settings, "COMPANY_NAME", "AG Construction")
    adres = getattr(settings, "COMPANY_ADDRESS", "")
    telefon = getattr(settings, "COMPANY_PHONE", "")
    email = getattr(settings, "COMPANY_EMAIL", "")

    # Parse address
    street = adres
    locality = ""
    postal = ""
    if adres:
        parts = adres.split(",")
        street = parts[0].strip() if parts else adres
        if len(parts) > 1:
            rest = parts[1].strip().split(" ", 1)
            postal = rest[0] if len(rest) > 1 else ""
            locality = rest[1] if len(rest) > 1 else rest[0]

    publisher = {
        "@type": "Organization",
        "name": nazwa,
        "vatID": nip,
        "taxID": regon,
        "address": {
            "@type": "PostalAddress",
            "streetAddress": street,
            "addressLocality": locality,
            "postalCode": postal,
            "addressCountry": "Polska",
        },
        "telephone": telefon,
        "email": email,
    }

    items = []
    for o in _get_oferty():
        cena = _aktualna_cena(o)
        cena_m2 = _cena_m2(o)
        inv = o.inwestycja

        additional = []
        if o.metraz:
            additional.append({"@type": "PropertyValue", "name": "Powierzchnia uzytkowa", "value": float(o.metraz), "unitText": "m2"})
        if o.pokoje:
            additional.append({"@type": "PropertyValue", "name": "Liczba pokoi", "value": o.pokoje})
        if cena_m2:
            additional.append({"@type": "PropertyValue", "name": "Cena za m2", "value": cena_m2, "unitText": "PLN/m2"})

        item = {
            "@type": "Product",
            "name": f"Lokal {o.numer_lokalu or o.pk}",
            "description": f"Lokal {o.numer_lokalu}, {o.pokoje} pok., {o.metraz} m²",
            "category": "Nieruchomosci/Lokale mieszkalne",
            "offers": {
                "@type": "Offer",
                "priceCurrency": "PLN",
                "price": float(cena) if cena else None,
                "availability": "https://schema.org/InStock",
                "seller": {"@type": "Organization", "name": nazwa, "vatID": nip},
            },
            "additionalProperty": additional,
            "itemOffered": {
                "@type": "Apartment",
                "address": {
                    "@type": "PostalAddress",
                    "streetAddress": o.adres,
                    "addressCountry": "Polska",
                },
                "numberOfRooms": o.pokoje,
                "floorSize": {"@type": "QuantitativeValue", "value": float(o.metraz) if o.metraz else None, "unitCode": "MTK"},
            },
            "pomieszczenia_przynalezne": [
                {"nazwa": p.nazwa, "powierzchnia_m2": float(p.powierzchnia) if p.powierzchnia else None, "cena_pln": float(p.cena) if p.cena else None}
                for p in o.pomieszczenia.all()
            ],
            "rabaty": [
                {"nazwa": r.nazwa, "wartosc": float(r.wartosc), "typ": r.typ, "data_od": str(r.data_od), "data_do": str(r.data_do)}
                for r in o.rabaty.all()
            ],
            "inne_swiadczenia_pieniezne": [
                {"nazwa": s.nazwa, "kwota_pln": float(s.kwota)}
                for s in o.swiadczenia.all()
            ],
            "id_przedsiewziecia": inv.unikalny_identyfikator_przedsiewziecia or f"inv-{inv.pk}",
        }
        items.append(item)

    return {
        "@context": {
            "@vocab": "http://schema.org/",
            "nip": "http://schema.org/vatID",
            "regon": "http://schema.org/taxID",
        },
        "@type": "Dataset",
        "name": f"Ceny ofertowe lokali dewelopera {nazwa}",
        "description": "Zbiór danych zawierający aktualne ceny ofertowe lokali mieszkalnych zgodnie z art. 19a i 19b ustawy o ochronie praw nabywcy lokalu mieszkalnego.",
        "dateModified": date.today().isoformat(),
        "license": "https://creativecommons.org/publicdomain/zero/1.0/",
        "publisher": publisher,
        "distribution": [
            {"@type": "DataDownload", "encodingFormat": "text/csv", "contentUrl": f"https://{getattr(settings, 'SITE_DOMAIN', 'localhost')}/api/data.csv"},
            {"@type": "DataDownload", "encodingFormat": "application/ld+json", "contentUrl": f"https://{getattr(settings, 'SITE_DOMAIN', 'localhost')}/api/data.jsonld"},
            {"@type": "DataDownload", "encodingFormat": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "contentUrl": f"https://{getattr(settings, 'SITE_DOMAIN', 'localhost')}/api/data.xlsx"},
        ],
        "itemListElement": items,
    }


def view_jsonld(request):
    data_dict = generate_jsonld()
    data_str = json.dumps(data_dict, ensure_ascii=False, indent=2)
    data_bytes = data_str.encode("utf-8")
    etag = _md5(data_bytes)
    today = date.today().strftime("%Y%m%d")
    resp = HttpResponse(data_bytes, content_type="application/ld+json; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="ag_construction_oferty_{today}.jsonld"'
    return _open_data_headers(resp, etag)


# ── METADATA XML ──────────────────────────────────────────────────────────────

def generate_metadata_xml() -> bytes:
    nazwa = getattr(settings, "COMPANY_NAME", "AG Construction")
    now = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
    today = date.today().isoformat()
    domain = getattr(settings, "SITE_DOMAIN", "localhost")
    base_url = f"https://{domain}/api"

    xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<datasets xmlns="urn:otwarte-dane:harvester:1.13">
  <dataset status="published">
    <extIdent>ag-construction-ceny-lokali-01</extIdent>
    <title>
      <polish>Ceny ofertowe lokali mieszkalnych dewelopera {nazwa}</polish>
      <english>Real estate offer prices by developer {nazwa}</english>
    </title>
    <description>
      <polish>Zbior danych zawierajacy aktualne ceny ofertowe lokali mieszkalnych zgodnie z art. 19a i 19b ustawy o ochronie praw nabywcy lokalu mieszkalnego (Dz. U. 2025 poz. 758).</polish>
      <english>Dataset containing current offer prices of residential units pursuant to the Act on the Protection of the Rights of a Purchaser of a Residential Unit.</english>
    </description>
    <url>{base_url}/</url>
    <updateFrequency>daily</updateFrequency>
    <categories>
      <category>ECON</category>
    </categories>
    <conditions>
      <source>true</source>
      <modification>true</modification>
      <dbOrCopyrightedLicenseChosen>CC0 1.0</dbOrCopyrightedLicenseChosen>
    </conditions>
    <resources>
      <resource status="published">
        <extIdent>ag-construction-csv-01</extIdent>
        <url>{base_url}/data.csv</url>
        <title><polish>Ceny lokali - format CSV</polish></title>
        <description><polish>Plik CSV z cenami ofertowymi lokali mieszkalnych (separator srednik, kodowanie UTF-8 BOM).</polish></description>
        <availability>remote</availability>
        <dataDate>{today}</dataDate>
        <lastUpdateDate>{now}</lastUpdateDate>
        <hasDynamicData>true</hasDynamicData>
        <dataDateUpdatePeriod>daily</dataDateUpdatePeriod>
      </resource>
      <resource status="published">
        <extIdent>ag-construction-jsonld-01</extIdent>
        <url>{base_url}/data.jsonld</url>
        <title><polish>Ceny lokali - format JSON-LD</polish></title>
        <description><polish>Plik JSON-LD (Schema.org) z cenami ofertowymi lokali mieszkalnych.</polish></description>
        <availability>remote</availability>
        <dataDate>{today}</dataDate>
        <lastUpdateDate>{now}</lastUpdateDate>
        <hasDynamicData>true</hasDynamicData>
        <dataDateUpdatePeriod>daily</dataDateUpdatePeriod>
      </resource>
      <resource status="published">
        <extIdent>ag-construction-xlsx-01</extIdent>
        <url>{base_url}/data.xlsx</url>
        <title><polish>Ceny lokali - format XLSX</polish></title>
        <description><polish>Plik Excel (XLSX) z cenami ofertowymi lokali mieszkalnych.</polish></description>
        <availability>remote</availability>
        <dataDate>{today}</dataDate>
        <lastUpdateDate>{now}</lastUpdateDate>
        <hasDynamicData>true</hasDynamicData>
        <dataDateUpdatePeriod>daily</dataDateUpdatePeriod>
      </resource>
    </resources>
    <tags>
      <tag lang="pl">deweloper</tag>
      <tag lang="pl">nieruchomosci</tag>
      <tag lang="pl">ceny lokali</tag>
      <tag lang="pl">mieszkania</tag>
      <tag lang="en">real estate</tag>
      <tag lang="en">developer</tag>
    </tags>
    <lastUpdateDate>{now}</lastUpdateDate>
    <hasDynamicData>true</hasDynamicData>
  </dataset>
</datasets>"""
    return xml.encode("utf-8")


def view_metadata_xml(request):
    data = generate_metadata_xml()
    etag = _md5(data)
    resp = HttpResponse(data, content_type="application/xml; charset=utf-8")
    return _open_data_headers(resp, etag)


# ── MD5 endpoints ─────────────────────────────────────────────────────────────

def view_md5(request, fmt):
    generators = {
        "csv": generate_csv,
        "xlsx": generate_xlsx,
        "jsonld": lambda: json.dumps(generate_jsonld(), ensure_ascii=False).encode("utf-8"),
        "xml": generate_metadata_xml,
    }
    gen = generators.get(fmt)
    if not gen:
        return HttpResponse("Not found", status=404)
    data = gen()
    return HttpResponse(_md5(data), content_type="text/plain")
